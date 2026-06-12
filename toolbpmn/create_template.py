"""Genera la plantilla Excel estandarizada para la BPMN Tool."""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from pathlib import Path

CASSA_BLUE  = "004B9A"
CASSA_GREEN = "00A651"
HEADER_BG   = "004B9A"
SUBHDR_BG   = "D6E4F7"
EXAMPLE_BG  = "F8FBFF"
REQUIRED_BG = "FFF3CD"

COLUMNS = [
    # (header, ancho, descripcion, obligatorio)
    ("ID_Paso",           12, "Identificador único (ej: step_1)",               True),
    ("Nombre_Actividad",  30, "Nombre corto de la tarea (verbo + objeto)",       True),
    ("Tipo",              14, "tarea / decision / inicio / fin / subproceso",    True),
    ("Area_Responsable",  25, "Nombre del área o rol que ejecuta",               True),
    ("Descripcion",       40, "Qué se hace exactamente en este paso",            True),
    ("Tiempo_Estimado",   16, "Número entero (ej: 30)",                          False),
    ("Unidad_Tiempo",     14, "minutos / horas / dias",                          False),
    ("Instrucciones",     45, "Instrucciones detalladas, sistemas, normas",      False),
    ("Condicion",         35, "Solo para tipo=decision: criterio de la decisión",False),
    ("Siguiente_Paso",    25, "ID(s) del siguiente paso. Separar con coma",      True),
]

EXAMPLE_ROWS = [
    ("step_1", "Inicio del proceso",       "inicio",   "Solicitante",   "El proceso inicia cuando llega la solicitud", 0, "", "", "", "step_2"),
    ("step_2", "Recibir solicitud",        "tarea",    "Solicitante",   "El solicitante llena el formulario en SAP",   15, "minutos", "Transacción ME51N en SAP", "", "step_3"),
    ("step_3", "Revisar monto",            "decision", "Jefe de Área",  "Evaluar si el monto requiere aprobación adicional", 10, "minutos", "", "Monto menor a $5,000", "step_4,step_5"),
    ("step_4", "Aprobar solicitud",        "tarea",    "Jefe de Área",  "El jefe aprueba la solicitud en el sistema",  5,  "minutos", "Sistema de aprobaciones", "", "step_6"),
    ("step_5", "Solicitar documentación",  "tarea",    "Jefe de Área",  "Pedir documentación adicional al solicitante",30, "minutos", "", "", "step_6"),
    ("step_6", "Fin del proceso",          "fin",      "Solicitante",   "El proceso concluye",                         0,  "", "", "", ""),
]


def create_template(output_path: str):
    wb = openpyxl.Workbook()

    # ── Hoja 1: Plantilla ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Proceso"

    thin = Side(style="thin", color="C5CFE0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Fila 1: título
    ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
    c = ws["A1"]
    c.value = "PLANTILLA DE LEVANTAMIENTO DE PROCESOS — Grupo CASSA"
    c.font = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
    c.fill = PatternFill("solid", fgColor=CASSA_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Fila 2: instrucción
    ws.merge_cells(f"A2:{get_column_letter(len(COLUMNS))}2")
    c = ws["A2"]
    c.value = ("Completa una fila por cada paso del proceso. "
               "Columnas marcadas (*) son obligatorias. "
               "Guarda el archivo y cárgalo en la BPMN Tool.")
    c.font = Font(italic=True, color="444444", size=9, name="Calibri")
    c.fill = PatternFill("solid", fgColor="E8F0FA")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 22

    # Fila 3: descripción de columnas
    for col_i, (hdr, ancho, desc, req) in enumerate(COLUMNS, 1):
        c = ws.cell(row=3, column=col_i, value=desc)
        c.font = Font(italic=True, color="555555", size=8, name="Calibri")
        c.fill = PatternFill("solid", fgColor=SUBHDR_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[3].height = 32

    # Fila 4: encabezados
    for col_i, (hdr, ancho, desc, req) in enumerate(COLUMNS, 1):
        label = f"{hdr} *" if req else hdr
        c = ws.cell(row=4, column=col_i, value=label)
        c.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        c.fill = PatternFill("solid", fgColor="1A5276" if req else "2980B9")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
        ws.column_dimensions[get_column_letter(col_i)].width = ancho
    ws.row_dimensions[4].height = 22

    # Filas de ejemplo
    for row_i, row_data in enumerate(EXAMPLE_ROWS, 5):
        for col_i, val in enumerate(row_data, 1):
            c = ws.cell(row=row_i, column=col_i, value=val)
            c.font = Font(size=9, name="Calibri", italic=True, color="555555")
            c.fill = PatternFill("solid", fgColor=EXAMPLE_BG)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = border
        ws.row_dimensions[row_i].height = 18

    # Separador visual entre ejemplos y datos reales
    row_sep = 5 + len(EXAMPLE_ROWS)
    ws.merge_cells(f"A{row_sep}:{get_column_letter(len(COLUMNS))}{row_sep}")
    c = ws.cell(row=row_sep, column=1,
                value="↑ Filas de ejemplo — bórralas y escribe tu proceso abajo ↓")
    c.font = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
    c.fill = PatternFill("solid", fgColor=CASSA_GREEN)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row_sep].height = 18

    # 10 filas vacías para que el usuario llene
    for row_i in range(row_sep + 1, row_sep + 11):
        for col_i in range(1, len(COLUMNS) + 1):
            c = ws.cell(row=row_i, column=col_i, value="")
            c.border = border
            c.alignment = Alignment(vertical="center")
            c.font = Font(size=10, name="Calibri")
        ws.row_dimensions[row_i].height = 18

    ws.freeze_panes = "A5"

    # ── Hoja 2: Referencia ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Referencia")
    ws2["A1"] = "VALORES VÁLIDOS POR COLUMNA"
    ws2["A1"].font = Font(bold=True, color="FFFFFF", size=11)
    ws2["A1"].fill = PatternFill("solid", fgColor=CASSA_BLUE)
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 50

    refs = [
        ("Columna", "Valores aceptados"),
        ("Tipo",    "inicio | tarea | decision | fin | subproceso | evento"),
        ("Unidad_Tiempo", "minutos | horas | dias"),
        ("Siguiente_Paso", "ID del paso siguiente. Si hay más de uno (en decisiones) separa con coma: step_4,step_5"),
        ("Condicion", "Solo para tipo=decision. Ejemplo: Monto menor a $5,000"),
        ("Tiempo_Estimado", "Número entero positivo. 0 para inicio/fin"),
    ]
    for row_i, (col_a, col_b) in enumerate(refs, 2):
        ws2[f"A{row_i}"] = col_a
        ws2[f"B{row_i}"] = col_b
        ws2[f"A{row_i}"].font = Font(bold=(row_i == 2), size=9)
        ws2[f"B{row_i}"].font = Font(size=9)
        ws2[f"A{row_i}"].fill = PatternFill("solid", fgColor=SUBHDR_BG if row_i == 2 else "FFFFFF")

    wb.save(output_path)
    print(f"Plantilla creada: {output_path}")


if __name__ == "__main__":
    out = str(Path(__file__).parent / "assets" / "Plantilla_Proceso_CASSA.xlsx")
    create_template(out)
