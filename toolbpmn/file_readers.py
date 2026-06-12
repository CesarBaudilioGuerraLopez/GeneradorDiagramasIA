"""Parsers para los distintos formatos de entrada de documentos."""

import io
from pathlib import Path
from typing import Any

# ── Mapeo de sinónimos → campo estándar ──────────────────────────────────────
_COLUMN_ALIASES: dict[str, str] = {
    # ID
    "id_paso": "id", "id": "id", "paso": "id", "step": "id", "num": "id",
    "numero": "id", "número": "id", "nro": "id",
    # Nombre
    "nombre_actividad": "nombre", "nombre": "nombre", "actividad": "nombre",
    "tarea": "nombre", "task": "nombre", "activity": "nombre",
    "nombre_tarea": "nombre", "nombre_paso": "nombre",
    # Tipo
    "tipo": "tipo", "type": "tipo", "clase": "tipo",
    # Rol / Responsable
    "area_responsable": "rol", "rol": "rol", "responsable": "rol",
    "area": "rol", "área": "rol", "responsible": "rol", "role": "rol",
    "ejecutor": "rol", "quien": "rol", "quién": "rol",
    "departamento": "rol", "unidad": "rol",
    # Descripción
    "descripcion": "descripcion", "descripción": "descripcion",
    "description": "descripcion", "detalle": "descripcion",
    "que_hace": "descripcion", "qué_hace": "descripcion",
    "objetivo_paso": "descripcion",
    # Tiempo
    "tiempo_estimado": "tiempo_ejecucion", "tiempo": "tiempo_ejecucion",
    "duracion": "tiempo_ejecucion", "duración": "tiempo_ejecucion",
    "time": "tiempo_ejecucion", "duration": "tiempo_ejecucion",
    # Unidad
    "unidad_tiempo": "unidad_tiempo", "unidad": "unidad_tiempo",
    "unit": "unidad_tiempo",
    # Instrucciones
    "instrucciones": "documentacion", "instruccion": "documentacion",
    "notas": "documentacion", "notes": "documentacion",
    "documentacion": "documentacion", "documentación": "documentacion",
    "procedimiento": "documentacion",
    # Condición
    "condicion": "condicion", "condición": "condicion",
    "condition": "condicion", "criterio": "condicion",
    "regla": "condicion",
    # Siguiente
    "siguiente_paso": "siguiente", "siguiente": "siguiente",
    "next": "siguiente", "next_step": "siguiente",
    "flujo": "siguiente", "secuencia": "siguiente",
}


def _normalize_header(h: str) -> str:
    """Normaliza un encabezado de columna para comparación."""
    s = str(h).replace("*", "").replace("(", "").replace(")", "").strip()
    s = s.lower()
    s = (s.replace("á", "a").replace("é", "e").replace("í", "i")
          .replace("ó", "o").replace("ú", "u").replace("ñ", "n"))
    s = s.replace(" ", "_")
    return s.strip("_")  # quita underscores sobrantes al inicio/fin


def read_txt(file) -> str:
    if hasattr(file, "read"):
        raw = file.read()
        return raw.decode("utf-8", errors="replace")
    return Path(file).read_text(encoding="utf-8", errors="replace")


def read_docx(file) -> str:
    from docx import Document
    if hasattr(file, "read"):
        doc = Document(io.BytesIO(file.read()))
    else:
        doc = Document(file)

    parts = []
    for block in _iter_blocks(doc):
        text = block.strip()
        if text:
            parts.append(text)
    return "\n".join(parts)


def _iter_blocks(doc):
    from docx.oxml.ns import qn
    body = doc.element.body
    for child in body:
        tag = child.tag.split("}")[-1] if "}" in child.tag else child.tag
        if tag == "p":
            yield "".join(n.text or "" for n in child.iter()
                          if hasattr(n, "text") and n.text)
        elif tag == "tbl":
            for row in child.iter(qn("w:tr")):
                cells = []
                for cell in row.iter(qn("w:tc")):
                    cells.append("".join(
                        n.text or "" for n in cell.iter()
                        if hasattr(n, "text") and n.text
                    ).strip())
                yield " | ".join(cells)


def read_xlsx(file) -> tuple[str, dict]:
    """
    Lee un Excel y devuelve:
      - texto plano para enviar a Claude
      - metadata con las columnas detectadas y su mapeo
    """
    import openpyxl

    if hasattr(file, "read"):
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
    else:
        wb = openpyxl.load_workbook(file, data_only=True)

    # Intentar detectar la hoja principal (primera no vacía)
    ws = None
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() in ("referencia", "reference", "ayuda", "help"):
            continue
        ws = wb[sheet_name]
        break
    if ws is None:
        ws = wb.active

    # ── Detectar fila de encabezados ─────────────────────────────────────────
    header_row_idx, col_mapping, detected = _detect_headers(ws)

    if col_mapping:
        # ── Modo plantilla: parsear columnas conocidas ────────────────────────
        text, metadata = _parse_template(ws, header_row_idx, col_mapping, detected)
    else:
        # ── Modo genérico: texto plano de todas las hojas ─────────────────────
        parts = []
        for sheet_name in wb.sheetnames:
            sws = wb[sheet_name]
            parts.append(f"=== Hoja: {sheet_name} ===")
            for row in sws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                line = " | ".join(cells).strip(" |")
                if line.replace("|", "").strip():
                    parts.append(line)
        text = "\n".join(parts)
        metadata = {"modo": "generico", "columnas_detectadas": {}}

    return text, metadata


def _detect_headers(ws) -> tuple[int, dict[str, int], dict[str, str]]:
    """
    Busca la fila de encabezados en las primeras 10 filas.
    Retorna (fila_idx_base1, {campo_std: col_idx_base1}, {campo_std: header_original}).
    """
    for row_idx in range(1, min(11, ws.max_row + 1)):
        row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        # Contar cuántas celdas coinciden con aliases conocidos
        matches: dict[str, int]  = {}   # campo_std → col_idx
        originals: dict[str, str] = {}  # campo_std → header original
        for col_idx, val in enumerate(row_vals, 1):
            if val is None:
                continue
            norm = _normalize_header(str(val))
            if norm in _COLUMN_ALIASES:
                campo = _COLUMN_ALIASES[norm]
                if campo not in matches:  # primera ocurrencia gana
                    matches[campo] = col_idx
                    originals[campo] = str(val).strip()

        if len(matches) >= 3:  # mínimo 3 columnas reconocidas para considerarlo plantilla
            return row_idx, matches, originals

    return 1, {}, {}


def _parse_template(ws, header_row: int, col_map: dict[str, int],
                    originals: dict[str, str]) -> tuple[str, dict]:
    """Extrae filas de datos y las convierte en texto estructurado para Claude."""
    rows_text = []
    data_rows = []

    for row_idx in range(header_row + 1, ws.max_row + 1):
        row: dict[str, Any] = {}
        for campo, col_idx in col_map.items():
            val = ws.cell(row=row_idx, column=col_idx).value
            row[campo] = str(val).strip() if val is not None else ""

        # Saltar filas completamente vacías o separadores
        if not any(row.values()):
            continue
        nombre = row.get("nombre", "")
        if not nombre or nombre.startswith("↑") or nombre.startswith("↓"):
            continue

        # Construir línea de texto descriptiva
        parts = []
        if row.get("id"):
            parts.append(f"Paso {row['id']}")
        if nombre:
            parts.append(f"'{nombre}'")
        if row.get("tipo"):
            parts.append(f"(tipo: {row['tipo']})")
        if row.get("rol"):
            parts.append(f"— Responsable: {row['rol']}")
        if row.get("descripcion"):
            parts.append(f"— {row['descripcion']}")
        if row.get("tiempo_ejecucion") and row["tiempo_ejecucion"] not in ("0", ""):
            u = row.get("unidad_tiempo", "minutos")
            parts.append(f"[{row['tiempo_ejecucion']} {u}]")
        if row.get("documentacion"):
            parts.append(f"[Instrucciones: {row['documentacion']}]")
        if row.get("condicion"):
            parts.append(f"[Criterio: {row['condicion']}]")
        if row.get("siguiente"):
            parts.append(f"→ Siguiente: {row['siguiente']}")

        rows_text.append(" ".join(parts))
        data_rows.append(row)

    # Texto final para Claude
    text = (
        "El siguiente proceso fue descrito usando una plantilla estandarizada. "
        "Cada línea es un paso. Respeta los IDs, responsables y flujo de secuencia exactamente como están:\n\n"
        + "\n".join(rows_text)
    )

    metadata = {
        "modo": "plantilla",
        "filas_detectadas": len(data_rows),
        "columnas_detectadas": {campo: orig for campo, orig in originals.items()},
        "data_rows": data_rows,
    }
    return text, metadata


def extract_text(file, filename: str) -> tuple[str, dict]:
    """
    Despacha al parser correcto.
    Retorna (texto, metadata).
    metadata incluye info de columnas detectadas para Excel.
    """
    ext = Path(filename).suffix.lower()
    if ext in (".txt", ".md"):
        return read_txt(file), {}
    if ext == ".docx":
        return read_docx(file), {}
    if ext in (".xlsx", ".xls"):
        return read_xlsx(file)
    raise ValueError(f"Formato no soportado: {ext}")
